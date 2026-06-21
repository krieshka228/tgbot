"""Тесты ФИО (full_name) в уведомлении админу и Excel-отчёте."""

import io
from datetime import datetime
from types import SimpleNamespace

import openpyxl

from bot.excel_reports import build_monthly_report
from bot.utils import format_order_for_admin


def _order(full_name=None, user_full_name="ТГ Имя"):
    user = SimpleNamespace(id=1, username=None, full_name=user_full_name,
                           phone="+79990001122", address="ул. Пушкина")
    item = SimpleNamespace(product=SimpleNamespace(name="Платье"),
                           product_id=1, quantity=2, price_at_order=1000.0)
    return SimpleNamespace(id=10, user=user, items=[item], full_name=full_name,
                           delivery_address="ул. Пушкина", delivery_method="СДЭК",
                           total_amount=2000.0, created_at=datetime(2026, 6, 1))


def test_format_order_for_admin_shows_order_full_name():
    text = format_order_for_admin(_order(full_name="Иванов Иван"))
    assert "ФИО: Иванов Иван" in text


def test_format_order_for_admin_falls_back_to_profile_name():
    text = format_order_for_admin(_order(full_name=None, user_full_name="Профиль Имя"))
    assert "ФИО: Профиль Имя" in text


def test_excel_has_fio_column_and_value():
    data = build_monthly_report([_order(full_name="Петров Пётр")], users=[])
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["Заказы"]
    headers = [c.value for c in ws[1]]
    assert "ФИО" in headers
    fio_col = headers.index("ФИО") + 1
    assert ws.cell(row=2, column=fio_col).value == "Петров Пётр"


def test_excel_fio_empty_when_absent():
    # Нет ни order.full_name, ни профиля → ячейка пустая.
    data = build_monthly_report([_order(full_name=None, user_full_name=None)], users=[])
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["Заказы"]
    headers = [c.value for c in ws[1]]
    fio_col = headers.index("ФИО") + 1
    assert ws.cell(row=2, column=fio_col).value in (None, "")
