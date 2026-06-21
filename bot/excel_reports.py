"""
excel_reports.py — Генерация Excel-отчётов.
Функции:
  build_monthly_report(orders, users) — отчёт за месяц: листы «Заказы» и «Клиенты»
  build_clients_excel(users)           — полная база клиентов
"""

import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _thin_border() -> Border:
    side = Side(style="thin")
    return Border(left=side, right=side, top=side, bottom=side)


def _header_style(cell, fill_color: str = "2E75B6") -> None:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _thin_border()


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def build_monthly_report(orders: list, users: list) -> bytes:
    """Создаёт Excel с листами «Заказы» и «Клиенты». В заказах отдельно Username и ФИО."""
    wb = openpyxl.Workbook()

    # ── Лист 1: Заказы ─────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Заказы"
    headers1 = [
        "№ заказа",
        "Username",
        "ФИО",
        "Телефон",
        "Способ доставки",
        "Адрес",
        "Товар",
        "Кол-во",
        "Цена, ₽",
        "Сумма, ₽",
        "Дата заказа",
    ]
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        _header_style(cell)

    row = 2
    for order in orders:
        user = order.user
        username = f"@{user.username}" if user and user.username else ""
        fio = getattr(order, "full_name", None) or (user.full_name if user else None) or ""
        phone = user.phone or "" if user else ""
        method = order.delivery_method or ""
        address = order.delivery_address or ""
        date_str = order.created_at.strftime("%d.%m.%Y") if order.created_at else ""

        if not order.items:
            ws1.cell(row=row, column=1, value=order.id).border = _thin_border()
            ws1.cell(row=row, column=2, value=username).border = _thin_border()
            ws1.cell(row=row, column=3, value=fio).border = _thin_border()
            ws1.cell(row=row, column=4, value=phone).border = _thin_border()
            ws1.cell(row=row, column=5, value=method).border = _thin_border()
            ws1.cell(row=row, column=6, value=address).border = _thin_border()
            ws1.cell(row=row, column=7, value="(нет позиций)").border = _thin_border()
            ws1.cell(row=row, column=8, value=0).border = _thin_border()
            ws1.cell(row=row, column=9, value=0).border = _thin_border()
            ws1.cell(row=row, column=10, value=0).border = _thin_border()
            ws1.cell(row=row, column=11, value=date_str).border = _thin_border()
            row += 1
        else:
            for item in order.items:
                name = item.product.name if item.product else f"Товар #{item.product_id}"
                subtotal = item.quantity * item.price_at_order
                ws1.cell(row=row, column=1, value=order.id).border = _thin_border()
                ws1.cell(row=row, column=2, value=username).border = _thin_border()
                ws1.cell(row=row, column=3, value=fio).border = _thin_border()
                ws1.cell(row=row, column=4, value=phone).border = _thin_border()
                ws1.cell(row=row, column=5, value=method).border = _thin_border()
                ws1.cell(row=row, column=6, value=address).border = _thin_border()
                ws1.cell(row=row, column=7, value=name).border = _thin_border()
                ws1.cell(row=row, column=8, value=item.quantity).border = _thin_border()
                ws1.cell(row=row, column=9, value=round(item.price_at_order, 2)).border = _thin_border()
                ws1.cell(row=row, column=10, value=round(subtotal, 2)).border = _thin_border()
                ws1.cell(row=row, column=11, value=date_str).border = _thin_border()
                row += 1

    _auto_width(ws1)

    # ── Лист 2: Клиенты ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Клиенты")
    headers2 = ["ID", "Имя (username)", "Имя (fullname)", "Телефон", "Адрес", "Согласие ПД", "Дата регистрации"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        _header_style(cell)

    row = 2
    for user in users:
        ws2.cell(row=row, column=1, value=user.id).border = _thin_border()
        ws2.cell(row=row, column=2, value=f"@{user.username}" if user.username else "").border = _thin_border()
        ws2.cell(row=row, column=3, value=user.full_name or "").border = _thin_border()
        ws2.cell(row=row, column=4, value=user.phone or "").border = _thin_border()
        ws2.cell(row=row, column=5, value=user.address or "").border = _thin_border()
        ws2.cell(row=row, column=6, value="Да" if user.consented else "Нет").border = _thin_border()
        date_str = user.created_at.strftime("%d.%m.%Y") if user.created_at else ""
        ws2.cell(row=row, column=7, value=date_str).border = _thin_border()
        row += 1

    _auto_width(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_clients_excel(users: list) -> bytes:
    """Создаёт Excel с полной базой клиентов. Приоритет: @username."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Клиенты"

    headers = ["ID", "Имя (username)", "Имя (fullname)", "Телефон", "Адрес", "Согласие ПД", "Дата регистрации"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        _header_style(cell)

    row = 2
    for user in users:
        ws.cell(row=row, column=1, value=user.id).border = _thin_border()
        ws.cell(row=row, column=2, value=f"@{user.username}" if user.username else "—").border = _thin_border()
        ws.cell(row=row, column=3, value=user.full_name or "—").border = _thin_border()
        ws.cell(row=row, column=4, value=user.phone or "—").border = _thin_border()
        ws.cell(row=row, column=5, value=user.address or "—").border = _thin_border()
        ws.cell(row=row, column=6, value="Да" if user.consented else "Нет").border = _thin_border()
        date_str = user.created_at.strftime("%d.%m.%Y") if user.created_at else "—"
        ws.cell(row=row, column=7, value=date_str).border = _thin_border()
        row += 1

    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()